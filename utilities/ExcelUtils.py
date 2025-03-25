from openpyxl import load_workbook
import pytest
from selenium import webdriver

class ExcelUtils:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.sheet = self.wb.active  # Assume data is in the first sheet

    def read_test_case(self, row):
        return self.sheet.cell(row=row, column=2).value  # Description column

    def update_status(self, row, status):
        self.sheet.cell(row=row, column=4, value=status)  # Update Status column
        self.wb.save(self.file_path)
